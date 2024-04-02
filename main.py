import openpyxl
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime

class Factura(object):
    def __init__(self, fechaEmitida, tipoFactura, tipoDoc, numeroDoc, medioPago, tipoTarjeta, precio, estado):
        self.fechaEmitida = fechaEmitida
        self.tipoFactura = tipoFactura
        self.tipoDoc = tipoDoc
        self.numeroDoc = numeroDoc
        self.medioPago = medioPago
        self.tipoTarjeta = tipoTarjeta
        self.precio = precio
        self.estado = estado


class Lector:
    def __init__(self):
        self.name = "./facturacion/facturacion.xlsx"  
        self.workbook = openpyxl.load_workbook(self.name)
        self.sheet = self.workbook.active
        self.current_row = 2
    
    def leer_factura(self):
        while self.current_row <= self.sheet.max_row:
            row = self.sheet[self.current_row]
            fecha_cell = row[0]  # Assuming fecha is in the first column of the row
            if isinstance(fecha_cell.value, datetime):
                fecha_str = fecha_cell.value.strftime("%d/%m/%Y")  # Intercambiar día y mes
            else:
                fecha_str = fecha_cell.value
            
            factura = Factura(fecha_str, *[cell.value for cell in row[1:]])  
            self.current_row += 1
            if factura.estado != 'PASO':
                return factura
        return None
        
    def actualizar_estado(self, row_number):
        estado_column = 8  # Assuming 'estado' column is the 8th column (adjust accordingly if it's different)
        self.sheet.cell(row=row_number, column=estado_column, value='PASO')
        self.workbook.save(self.name)
        
class Emisor(Factura, Lector):
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.url = "https://auth.afip.gob.ar/contribuyente_/login.xhtml?action=SYSTEM&system=rcel"
        self.driver = webdriver.Chrome()
        self.driver.get(self.url)
        self.login()

        self.emitir()    
    
    def login(self):
        # Encontrar y completar el campo de CUIT/CUIL
        username_input = self.driver.find_element(By.XPATH, '//*[@id="F1:username"]')
        username_input.send_keys(self.username)

        # Hacer clic en el botón Siguiente
        siguiente_button = self.driver.find_element(By.XPATH, '//*[@id="F1:btnSiguiente"]')
        siguiente_button.click()

        # Esperar a que se cargue la página
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="F1:password"]')))

        # Encontrar y completar el campo de Clave Fiscal
        password_input = self.driver.find_element(By.XPATH, '//*[@id="F1:password"]')
        password_input.send_keys(self.password)

        # Hacer clic en el botón Ingresar
        ingresar_button = self.driver.find_element(By.XPATH, '//*[@id="F1:btnIngresar"]')
        ingresar_button.click()

    # Lee y emite una factura
    def emitir(self):
        lectorFactura = Lector() # Creamos una instancia del lector de facturas
        while True:
            factura = lectorFactura.leer_factura() # Lee la factura
            if factura is None:
                break  # No hay más facturas por leer, terminar el proceso
            if factura.estado != 'PASO': # Verifica si la factura ya fue emitida
                # Seleccionar empresa
                empresa = self.driver.find_element(By.XPATH, '//*[@id="contenido"]/form/table/tbody/tr[4]/td/input[2]').click()
                 
                # Generar Comprobantes
                generarComprobantes = self.driver.find_element(By.XPATH, '//*[@id="btn_gen_cmp"]/span[2]').click()

                # Punto de venta
                puntoVenta = self.driver.find_element(By.XPATH, '//*[@id="puntodeventa"]').click()
                seleccionPunto = self.driver.find_element(By.XPATH, '//*[@id="puntodeventa"]/option[2]').click()

                # Continuar
                continuarBoton = self.driver.find_element(By.XPATH, '//*[@id="contenido"]/form/input[2]').click()
                
                # Ingresamos fecha de factura
                fechaComprobante = self.driver.find_element(By.XPATH, '//*[@id="fc"]').clear()
                fechaComprobante = self.driver.find_element(By.XPATH, '//*[@id="fc"]').send_keys(factura.fechaEmitida)

                # Elige el concepto, en este caso sera producto
                concepto = self.driver.find_element(By.XPATH, '//*[@id="idconcepto"]').click()
                conceptoProducto = self.driver.find_element(By.XPATH, '//*[@id="idconcepto"]/option[2]').click()

                # Continuar
                continuar = self.driver.find_element(By.XPATH, '//*[@id="contenido"]/form/input[2]').click() 

                # Elige tipo de consumidor, en este caso sera consumidor final
                consumidor = self.driver.find_element(By.XPATH, '//*[@id="idivareceptor"]').click()
                consumidorFinal = self.driver.find_element(By.XPATH, '//*[@id="idivareceptor"]/option[4]').click()

                # Elige el tipo de documento segun el lector
                documento = self.driver.find_element(By.XPATH, '//*[@id="idtipodocreceptor"]').click()
                if factura.tipoDoc == "DNI":
                    dniDocumento = self.driver.find_element(By.XPATH, '//*[@id="idtipodocreceptor"]/option[7]').click()
                elif factura.tipoDoc == "CUIL":
                    cuilDocumento = self.driver.find_element(By.XPATH, '//*[@id="idtipodocreceptor"]/option[2]').click()
                elif factura.tipoDoc == "CUIT":
                    cuitDocumento = self.driver.find_element(By.XPATH, '//*[@id="idtipodocreceptor"]/option[1]').click()
                
                # Introducir numero del documento seleccionado
                documentoNum = self.driver.find_element(By.XPATH, '//*[@id="nrodocreceptor"]')
                documentoNum.send_keys(factura.numeroDoc)
                documentoNum.send_keys(Keys.ENTER)

                # Elige medio de pago tarjeta o qr segun factura
                if factura.medioPago != 'QR' and factura.medioPago != 'OTRO':
                    if factura.medioPago == 'DEBITO':
                        debitoMedio = self.driver.find_element(By.XPATH, '//*[@id="formadepago2"]').click()
                        self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_debito1"]').click()
                        if factura.tipoTarjeta == 'MASTERCARD':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_debito1"]/option[2]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_debito1"]').send_keys('11111111111111111111')
                        elif factura.tipoTarjeta == 'VISA':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_debito1"]/option[3]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_debito1"]').send_keys('11111111111111111111')
                        elif factura.tipoTarjeta == 'CABAL':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_debito1"]/option[4]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_debito1"]').send_keys('11111111111111111111')
                        else:
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_debito1"]/option[5]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_debito1"]').send_keys('11111111111111111111')
                    elif factura.medioPago == 'CREDITO':
                        self.driver.find_element(By.XPATH, '//*[@id="formadepago3"]').click()
                        self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_credito1"]').click()
                        if factura.tipoTarjeta == 'VISA':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_credito1"]/option[3]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_credito1"]').send_keys('11111111111111111111')
                        elif factura.tipoTarjeta == 'MASTERCARD':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_credito1"]/option[4]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_credito1"]').send_keys('11111111111111111111')
                        elif factura.tipoTarjeta == 'CABAL':
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_id_tipo_credito1"]/option[7]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="tarjeta_nro_credito1"]').send_keys('11111111111111111111')                
                
                else:
                    self.driver.find_element(By.XPATH, '//*[@id="formadepago7"]').click()

                # Continuar 
                self.driver.find_element(By.XPATH, '//*[@id="formulario"]/input[2]').click()

                # Producto
                self.driver.find_element(By.XPATH, '//*[@id="detalle_descripcion1"]').send_keys('varios')

                # Precio Unitario
                self.driver.find_element(By.XPATH, '//*[@id="detalle_precio1"]').send_keys(factura.precio)

                # Actualizar estado en el archivo de facturacion 
                lectorFactura.actualizar_estado(lectorFactura.current_row - 1)
                
            # Esperar antes de procesar la siguiente factura
            time.sleep(60)

# Emisor de factura
e = Emisor("username", "password")
